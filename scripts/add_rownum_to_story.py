import openpyxl as px
import sys
import json
import re
import os
import random


xlsx_file = sys.argv[1]


wb = px.load_workbook(xlsx_file)
	
for sheet in wb:
	row_num = 0
	for row in sheet.iter_rows():
		row_num+=1
		if(row_num == 1):
			sheet.cell(row=row_num, column=9).value = 'DO NOT CHANGE'
		else:
			sheet.cell(row=row_num, column=9).value = sheet.title + '_' + str(row_num)			

wb.save(xlsx_file)