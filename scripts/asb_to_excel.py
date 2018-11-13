import openpyxl as px
import sys
from bs4 import BeautifulSoup
from urllib import request
import re
import os

def parse_story(soup, lang):
    row = 1
    ws_list = [
    {
        'type': 'type',
        'header': 'header',
        'title': 'title_'+lang,
        'content': 'content_'+lang,
        'option': 'option'
    }
    ]
    row = row+1
    title_img_el = soup.find('div',class_='cover-image')['style']
    m = re.search(r"\(.*\)", title_img_el)
    title_img = m.group(0)[1:-1]
    img_name = title_img.split('/')[-1]
    title = soup.find('div',class_='cover_title').text
    content = soup.find('div',class_='cover_author').text
    content += '\n'+soup.find('div',class_='cover_artist').text
    if not os.path.isfile('asb/'+img_name):
        url_get("https://www.africanstorybook.org/"+title_img, 'asb/'+img_name)

    # print(title_img,',',title,',',content)
    ws_list.append({
        'type': 'topic',
        'header': 'asb/'+img_name,
        'title': title,
        'content': content,
        'option': ''
    })
    row = row+1

    for tag in soup.find_all('div',class_='swiper-slide'):
        img = ''
        text = ''
        img_name = ''
        img_el = tag.find('div',class_='page-image')
        if img_el:
            m = re.search(r"\(.*\)", img_el['style'])
            img = m.group(0)[1:-1]
        text_el = tag.find('div',class_='page-text-story')
        if not text_el:
            text_el = tag.find('div',class_='page-textonly-story')
        if text_el:
            text = text_el.get_text("\n", strip=True)
        if img != '' or text != '':
            if img != '':
                img_name = img.split('/')[-1]
                if not os.path.isfile('asb/'+img_name):
                    url_get("https://www.africanstorybook.org/"+img, 'asb/'+img_name)

            # print(img, ',', text)
            row_dict = {
                'type': 'article'
            }
            if img_name:
                row_dict['header'] = 'asb/'+img_name
            else:
                row_dict['header'] = ''
            row_dict['title'] = ''
            row_dict['content'] = text
            row_dict['option'] = ''
            ws_list.append(row_dict)
            row = row+1
    back_cover = soup.find('div',class_='backcover_wrapper')
    if back_cover:
        back_cover_text = back_cover.get_text("\n", strip=True)
        # print(back_cover_text)
        ws_list.append({
            'type': 'article',
            'header': '',
            'title': '',
            'content': back_cover_text,
            'option': ''
        })
    return ws_list


def url_get(url, file):
    print("downloading "+ url)
    remaining_download_tries = 15
    while remaining_download_tries > 0:
        try:
            s = request.urlretrieve(url, file)
        except Exception as e:
            print("error downloading " + url +" on trial no: " + str(15 - remaining_download_tries))
            remaining_download_tries = remaining_download_tries - 1
            continue
        else:
            break

story_file = sys.argv[1]
story_name = os.path.splitext(story_file)[0]

wb = px.load_workbook(story_name+'.xlsx')
worksheets = [s.title for s in wb.worksheets];
with open(story_file) as f:
    for story_id in f:
        story_id = story_id.rstrip()
        print(story_id)
        ws = wb.active
        if story_id in worksheets:
            # ws = wb[story_id]
            continue
        else:
            ws = wb.create_sheet(story_id)
        with request.urlopen('https://www.africanstorybook.org/read/readbook.php?id='+story_id+'&d=0&a=1') as f:
            soup = BeautifulSoup(f, features="html.parser")
            try:
                s = soup.find('div',class_='list-bliock').find(string=re.compile('English'))
                if s == None:
                    print('Skipping '+story_id)
                    continue
            except Exception as e:
                print('Skipping '+story_id)
                continue

            sw_list = parse_story(soup, 'sw')

            try:
                s = soup.find('div',class_='list-bliock').find(string=re.compile('English'))

                if s:
                    link = s.parent.parent.parent.parent['onclick']
                    m = re.search(r"\d+",link)
                    eng_story_id = m.group(0)
                    with request.urlopen('https://www.africanstorybook.org/read/readbook.php?id='+eng_story_id+'&d=0&a=1') as eng_f:
                        soup = BeautifulSoup(eng_f, features="html.parser")
                        en_list = parse_story(soup, 'en')
                        eng_swa_same = False
                        if len(sw_list) == len(en_list):
                            eng_swa_same = True
                            for i in range(len(en_list)):
                                en_row = en_list[i]
                                sw_row = sw_list[i]
                                if sw_row['type'] == en_row['type'] and sw_row['header'] == en_row['header']:
                                    sw_list[i]['title_en'] = en_row['title']
                                    sw_list[i]['content_en'] = en_row['content']
                                else:
                                    print('Skipping',en_row, sw_row)
                                    eng_swa_same = False
                        if not eng_swa_same:
                            print('Skipping',en_list, sw_list)
                            continue
                row = 1
                for row_dict in sw_list:
                    # print(row_dict)
                    ws.cell(row=row, column=1).value = row_dict['type']
                    ws.cell(row=row, column=2).value = row_dict['header']
                    ws.cell(row=row, column=3).value = row_dict['title_en']
                    ws.cell(row=row, column=4).value = row_dict['title']
                    ws.cell(row=row, column=5).value = row_dict['content_en']
                    ws.cell(row=row, column=6).value = row_dict['content']
                    ws.cell(row=row, column=7).value = row_dict['option']
                    row = row+1
            except Exception as e:
                print('Error downloading: ' +story_id + e)
        wb.save(story_name+'.xlsx')
